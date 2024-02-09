
from django.shortcuts import render,redirect
from Register_Login.models import *
from Register_Login.views import logout

# Create your views here.
from Company_Staff.models import Vendor, Vendor_comments_table, Vendor_doc_upload_table, Vendor_mail_table,Vendor_remarks_table,VendorContactPerson,VendorHistory
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from openpyxl import load_workbook
from django.http import HttpResponseBadRequest, HttpResponseNotFound, JsonResponse
import os
from datetime import date
from email.message import EmailMessage
from io import BytesIO
from django.conf import settings
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.shortcuts import get_object_or_404
from django.core.mail import EmailMessage
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


# -------------------------------Company section--------------------------------

# company dashboard
def company_dashboard(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_dash.html', context)
    else:
        return redirect('/')


# company profile
def company_profile(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_profile.html', context)
    else:
        return redirect('/')

# company profile
def company_staff_request(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        staff_request=StaffDetails.objects.filter(company=dash_details.id, company_approval=0).order_by('-id')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'requests':staff_request,
        }
        return render(request, 'company/staff_request.html', context)
    else:
        return redirect('/')

# company staff accept or reject
def staff_request_accept(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=1
    staff.save()
    return redirect('company_staff_request')

def staff_request_reject(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    login_details=LoginDetails.objects.get(id=staff.company.id)
    login_details.delete()
    staff.delete()
    return redirect('company_staff_request')


# All company staff view, cancel staff approval
def company_all_staff(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        all_staffs=StaffDetails.objects.filter(company=dash_details.id, company_approval=1).order_by('-id')
        print(all_staffs)
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'staffs':all_staffs,
        }
        return render(request, 'company/all_staff_view.html', context)
    else:
        return redirect('/')

def staff_approval_cancel(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=2
    staff.save()
    return redirect('company_all_staff')









# -------------------------------Staff section--------------------------------

# staff dashboard
def staff_dashboard(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_dash.html',context)
    else:
        return redirect('/')


# staff profile
def staff_profile(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_profile.html',context)
    else:
        return redirect('/')











# -------------------------------Zoho Modules section--------------------------------
 
 
 #-------------------------------Arya E.R-----------------------------------------------

 ##### Vendor #####
    
def vendor(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
    
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        comp_payment_terms=Company_Payment_Term.objects.filter(company=dash_details)
        if log_details.user_type=='Staff':

            return render(request,'zohomodules/vendor/create_vendor.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details}) 
        else:
            return render(request,'zohomodules/vendor/create_vendor.html',{'details':dash_details,'allmodules': allmodules,'comp_payment_terms':comp_payment_terms,'log_details':log_details}) 
    else:
        return redirect('/')

def view_vendor_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)

        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)

        allmodules= ZohoModules.objects.get(company=dash_details,status='New')  

        data=Vendor.objects.filter(company=dash_details)

         # Pagination
        
        # page = request.GET.get('page', 1)
        # paginator = Paginator(data, 5)

        # try:
        #     items = paginator.page(page)
        # except PageNotAnInteger:
        #     items = paginator.page(1)
        # except EmptyPage:
        #     items = paginator.page(paginator.num_pages)

        return render(request,'zohomodules/vendor/vendor_list.html',{'details':dash_details,'allmodules': allmodules,'data':data,'log_details':log_details}) 


    else:
        return redirect('/')


# @login_required(login_url='login')
def add_vendor(request):
   
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)

        

       
        if request.method=="POST":
            vendor_data=Vendor()
            vendor_data.login_details=log_details
            vendor_data.company=dash_details
            vendor_data.title = request.POST.get('salutation')
            vendor_data.first_name=request.POST['first_name']
            vendor_data.last_name=request.POST['last_name']
            vendor_data.company_name=request.POST['company_name']
            vendor_data.vendor_display_name=request.POST['v_display_name']
            vendor_data.vendor_email=request.POST['vendor_email']
            vendor_data.phone=request.POST['w_phone']
            vendor_data.mobile=request.POST['m_phone']
            vendor_data.skype_name_number=request.POST['skype_number']
            vendor_data.designation=request.POST['designation']
            vendor_data.department=request.POST['department']
            vendor_data.website=request.POST['website']
            vendor_data.gst_treatment=request.POST['gst']
            vendor_data.vendor_status="Active"

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                vendor_data.pan_number=request.POST['pan_number']
                vendor_data.gst_number="null"
            else:
                vendor_data.gst_number=request.POST['gst_number']
                vendor_data.pan_number=request.POST['pan_number']

            vendor_data.source_of_supply=request.POST['source_supply']
            vendor_data.currency=request.POST['currency']
            vendor_data.opening_balance_type=request.POST['op_type']
            vendor_data.opening_balance=request.POST['opening_bal']
            vendor_data.payment_term=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])

           
            vendor_data.billing_attention=request.POST['battention']
            vendor_data.billing_country=request.POST['bcountry']
            vendor_data.billing_address=request.POST['baddress']
            vendor_data.billing_city=request.POST['bcity']
            vendor_data.billing_state=request.POST['bstate']
            vendor_data.billing_pin_code=request.POST['bzip']
            vendor_data.billing_phone=request.POST['bphone']
            vendor_data.billing_fax=request.POST['bfax']
            vendor_data.shipping_attention=request.POST['sattention']
            vendor_data.shipping_country=request.POST['s_country']
            vendor_data.shipping_address=request.POST['saddress']
            vendor_data.shipping_city=request.POST['scity']
            vendor_data.shipping_state=request.POST['sstate']
            vendor_data.shipping_pin_code=request.POST['szip']
            vendor_data.shipping_phone=request.POST['sphone']
            vendor_data.shipping_fax=request.POST['sfax']
            vendor_data.save()
           # ................ Adding to History table...........................
            
            vendor_history_obj=VendorHistory()
            vendor_history_obj.company=dash_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.vendor=vendor_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Registration Completed'
            vendor_history_obj.save()

    # .......................................................adding to remaks table.....................
            vdata=Vendor.objects.get(id=vendor_data.id)
            vendor=vdata
            rdata=Vendor_remarks_table()
            rdata.remarks=request.POST['remark']
            rdata.company=dash_details
            rdata.vendor=vdata
            rdata.save()


     #...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            vdata=Vendor.objects.get(id=vendor_data.id)
            vendor=vdata
           
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department)
                    mapped2=list(mapped2)
                    print(mapped2)
                    for ele in mapped2:
                        created = VendorContactPerson.objects.get_or_create(title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype_name_number=ele[6],designation=ele[7],department=ele[8],company=dash_details,vendor=vendor)
                
        
            messages.success(request, 'Data saved successfully!')   

            return redirect('view_vendor_list')
        
        else:
            messages.error(request, 'Some error occurred !')   

            return redirect('view_vendor_list')
    


def sort_vendor_by_name(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
  
        data=Vendor.objects.filter(login_details=log_details).order_by('first_name')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
            return redirect('/')    

def sort_vendor_by_amount(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
   
        data=Vendor.objects.filter(login_details=log_details).order_by('opening_balance')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
         return redirect('/') 

def view_vendor_active(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
   
        data=Vendor.objects.filter(login_details=log_details,vendor_status='Active').order_by('-id')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
         return redirect('/') 

    
    
def view_vendor_inactive(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
   
        data=Vendor.objects.filter(login_details=log_details,vendor_status='Inactive').order_by('-id')
        return render(request,'zohomodules/vendor/vendor_list.html',{'data':data,'dash_details':dash_details})
    else:
         return redirect('/') 
    
def delete_vendor(request,pk):
    if Vendor_comments_table.objects.filter(vendor=pk).exists():
        user2=Vendor_comments_table.objects.filter(vendor=pk)
        user2.delete()
    if Vendor_mail_table.objects.filter(vendor=pk).exists():
        user3=Vendor_mail_table.objects.filter(vendor=pk)
        user3.delete()
    if Vendor_doc_upload_table.objects.filter(vendor=pk).exists():
        user4=Vendor_doc_upload_table.objects.filter(vendor=pk)
        user4.delete()
    if VendorContactPerson.objects.filter(vendor=pk).exists():
        user5=VendorContactPerson.objects.filter(vendor=pk)
        user5.delete()
    if Vendor_remarks_table.objects.filter(vendor=pk).exists():
        user6=Vendor_remarks_table.objects.filter(vendor=pk)
        user6.delete()
    
    user1=Vendor.objects.get(id=pk)
    user1.delete()
    return redirect("view_vendor_list")



def view_vendor_details(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        vendor_obj=Vendor.objects.get(id=pk)
        vendor_comments=Vendor_comments_table.objects.filter(vendor=vendor_obj)
        vendor_history=VendorHistory.objects.filter(vendor=vendor_obj)
    
    content = {
                'details': dash_details,
               
                'allmodules': allmodules,
                'vendor_obj':vendor_obj,
              'log_details':log_details,
              'vendor_comments':vendor_comments,
              'vendor_history':vendor_history,
        }
    return render(request,'zohomodules/vendor/vendor_detailsnew.html',content)

def import_vendor_excel(request):
   if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)

        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)
            

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        if request.method == 'POST' :
       
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(1, eb.max_row + 1):
                    vendorsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    Vendor_object=Vendor(title=vendorsheet[0],first_name=vendorsheet[1],last_name=vendorsheet[2],company_name=vendorsheet[3],vendor_email=vendorsheet[4],phone=vendorsheet[5],mobile=vendorsheet[6],skype_name_number=vendorsheet[7],designation=vendorsheet[8],department=vendorsheet[9],website=vendorsheet[10],
                                         gst_treatment=vendorsheet[11],source_of_supply=vendorsheet[12],currency=vendorsheet[13],opening_balance_type=vendorsheet[14],
                                         opening_balance=vendorsheet[15],payment_term=vendorsheet[16],billing_attention=vendorsheet[17],billing_address=vendorsheet[18],
                                         billing_city=vendorsheet[19],billing_state=vendorsheet[20],billing_country=vendorsheet[21],billing_pin_code=vendorsheet[22],
                                         billing_phone=vendorsheet[23],billing_fax=vendorsheet[24],shipping_attention=vendorsheet[25],shipping_address=vendorsheet[26],shipping_city=vendorsheet[27],
                                         shipping_state=vendorsheet[28],shipping_country=vendorsheet[29],shipping_pin_code=vendorsheet[30],
                                         shipping_phone=vendorsheet[31], shipping_fax=vendorsheet[32], remarks=vendorsheet[33],company=dash_details,login_details=log_details)
                    Vendor_object.save()

    
                   
                messages.warning(request,'file imported')
                return redirect('view_vendor_list')    

    
            messages.error(request,'File upload Failed!11')
            return redirect('view_vendor_list')
        else:
            messages.error(request,'File upload Failed!11')
            return redirect('view_vendor_list')
        
def Vendor_edit(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)

    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        vendor_obj=Vendor.objects.get(id=pk)

    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        vendor_obj=Vendor.objects.get(id=pk)
    vendor_contact_obj=VendorContactPerson.objects.filter(vendor=vendor_obj)  
    comp_payment_terms=Company_Payment_Term.objects.filter(company=dash_details)
   
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'vendor_obj':vendor_obj,
            'log_details':log_details,
            'vendor_contact_obj':vendor_contact_obj,
            'comp_payment_terms':comp_payment_terms,
    }
   

    return render(request,'zohomodules/vendor/edit_vendor.html',content)

def do_vendor_edit(request,pk):
     if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        if request.method=="POST":
            vendor_data=Vendor.objects.get(id=pk)
            vendor_data.login_details=log_details
            vendor_data.company=dash_details
            vendor_data.title = request.POST.get('salutation')
            vendor_data.first_name=request.POST['first_name']
            vendor_data.last_name=request.POST['last_name']
            vendor_data.company_name=request.POST['company_name']
            vendor_data.vendor_display_name=request.POST['v_display_name']
            vendor_data.vendor_email=request.POST['vendor_email']
            vendor_data.phone=request.POST['w_phone']
            vendor_data.mobile=request.POST['m_phone']
            vendor_data.skype_name_number=request.POST['skype_number']
            vendor_data.designation=request.POST['designation']
            vendor_data.department=request.POST['department']
            vendor_data.website=request.POST['website']
            vendor_data.gst_treatment=request.POST['gst']
            vendor_data.vendor_status="Active"

            x=request.POST['gst']
            if x=="Unregistered Business-not Registered under GST":
                vendor_data.pan_number=request.POST['pan_number']
                vendor_data.gst_number="null"
            else:
                vendor_data.gst_number=request.POST['gst_number']
                vendor_data.pan_number=request.POST['pan_number']

            vendor_data.source_of_supply=request.POST['source_supply']
            vendor_data.currency=request.POST['currency']
            vendor_data.opening_balance_type=request.POST['op_type']
            vendor_data.opening_balance=request.POST['opening_bal']
            vendor_data.payment_term=Company_Payment_Term.objects.get(id=request.POST['payment_terms'])
            
           
            vendor_data.billing_attention=request.POST['battention']
            vendor_data.billing_country=request.POST['bcountry']
            vendor_data.billing_address=request.POST['baddress']
            vendor_data.billing_city=request.POST['bcity']
            vendor_data.billing_state=request.POST['bstate']
            vendor_data.billing_pin_code=request.POST['bzip']
            vendor_data.billing_phone=request.POST['bphone']
            vendor_data.billing_fax=request.POST['bfax']
            vendor_data.shipping_attention=request.POST['sattention']
            vendor_data.shipping_country=request.POST['s_country']
            vendor_data.shipping_address=request.POST['saddress']
            vendor_data.shipping_city=request.POST['scity']
            vendor_data.shipping_state=request.POST['sstate']
            vendor_data.shipping_pin_code=request.POST['szip']
            vendor_data.shipping_phone=request.POST['sphone']
            vendor_data.shipping_fax=request.POST['sfax']
            vendor_data.save()


              # ................ Adding to History table...........................
            
            vendor_history_obj=VendorHistory()
            vendor_history_obj.company=dash_details
            vendor_history_obj.login_details=log_details
            vendor_history_obj.vendor=vendor_data
            vendor_history_obj.date=date.today()
            vendor_history_obj.action='Edited'
            vendor_history_obj.save()
    # .......................................................adding to remaks table.....................
            vdata=Vendor.objects.get(id=vendor_data.id)
            rdata=Vendor_remarks_table.objects.get(vendor=vdata)
            rdata.remarks=request.POST['remark']
            rdata.company=dash_details
            rdata.vendor=vdata
            rdata.save()


    #  ...........................adding multiple rows of table to model  ........................................................  
        
            title =request.POST.getlist('salutation[]')
            first_name =request.POST.getlist('first_name[]')
            last_name =request.POST.getlist('last_name[]')
            email =request.POST.getlist('email[]')
            work_phone =request.POST.getlist('wphone[]')
            mobile =request.POST.getlist('mobile[]')
            skype_name_number =request.POST.getlist('skype[]')
            designation =request.POST.getlist('designation[]')
            department =request.POST.getlist('department[]') 
            person = request.POST.getlist('contact_person_id[]')
            vdata=Vendor.objects.get(id=vendor_data.id)
            vendor=vdata
            print(person)
            if title != ['Select']:
                if len(title)==len(first_name)==len(last_name)==len(email)==len(work_phone)==len(mobile)==len(skype_name_number)==len(designation)==len(department)==len(person):
                    mapped2=zip(title,first_name,last_name,email,work_phone,mobile,skype_name_number,designation,department,person)
                    mapped2=list(mapped2)
                    for ele in mapped2:
                       
                        existing_instance = VendorContactPerson.objects.filter(id=ele[9], company=dash_details, vendor=vendor).first()
                        if existing_instance:
                            # Update the existing instance
                            existing_instance.title = ele[0]
                            existing_instance.first_name = ele[1]
                            existing_instance.last_name = ele[2]
                            existing_instance.email = ele[3]
                            existing_instance.work_phone  = ele[4]
                            existing_instance.mobile = ele[5]
                            existing_instance.skype_name_number = ele[6]
                            existing_instance.designation = ele[7]
                            existing_instance.department = ele[8]

                            # Update other fields

                            existing_instance.save()
                        else:
                            # Create a new instance
                            new_instance = VendorContactPerson.objects.create(
                                title=ele[0],first_name=ele[1],last_name=ele[2],email=ele[3],
                                work_phone=ele[4],mobile=ele[5],skype_name_number=ele[6],designation=ele[7],department=ele[8],company=dash_details,vendor=vendor
                            )
            return redirect('view_vendor_list')
    

def delete_vendors(request, pk):
    try:
        vendor_obj = Vendor.objects.get(id=pk)

        vendor_obj.delete()
        return redirect('view_vendor_list')  
    except Vendor.DoesNotExist:
        return HttpResponseNotFound("Vendor not found.")
    
def vendor_status(request,pk):
    vendor_obj = Vendor.objects.get(id=pk)
    if vendor_obj.vendor_status == 'Active':
        vendor_obj.vendor_status ='Inactive'
    elif vendor_obj.vendor_status == 'Inactive':
        vendor_obj.vendor_status ='Active'
    vendor_obj.save()
    return redirect('view_vendor_details',pk)   

def vendor_add_comment(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
  
        if request.method =='POST':
            comment_data=request.POST['comments']
       
            vendor_id= Vendor.objects.get(id=pk) 
            vendor_obj=Vendor_comments_table()
            vendor_obj.comment=comment_data
            vendor_obj.vendor=vendor_id
            vendor_obj.company=dash_details
            vendor_obj.login_details= LoginDetails.objects.get(id=log_id)

            vendor_obj.save()
            return redirect('view_vendor_details',pk)
    return redirect('view_vendor_details',pk) 


def vendor_delete_comment(request, pk):
    try:
        vendor_comment =Vendor_comments_table.objects.get(id=pk)
        vendor_id=vendor_comment.vendor.id
        vendor_comment.delete()
        return redirect('view_vendor_details',vendor_id)  
    except Vendor_comments_table.DoesNotExist:
        return HttpResponseNotFound("comments not found.")
    

def add_vendor_file(request,pk):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)
        if request.method == 'POST':
            data=request.FILES.getlist('file')
            try:
                for doc in data:

                    vendor_obj=Vendor_doc_upload_table()
                    
                    vendor_obj.document = doc
                    vendor_obj.login_details = log_details
                    vendor_obj.company = dash_details
                    vendor_obj.vendor = Vendor.objects.get(id=pk)
                    vendor_obj.save()
                
                messages.success(request,'File uploaded')
                return redirect('view_vendor_details',pk) 
            except Vendor_doc_upload_table.DoesNotExist:
                return redirect('view_vendor_details',pk) 


    
def vendor_shareemail(request,pk):
    try:
            if request.method == 'POST':
                emails_string = request.POST['email']

    
                emails_list = [email.strip() for email in emails_string.split(',')]
                vendor_obj=Vendor.objects.get(id=pk)
                        
                context = {'vendor_obj':vendor_obj}
                template_path = 'vendor/vendormailoverview.html'
                template = get_template(template_path)
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'{vendor_obj.first_name}details - {vendor_obj.id}.pdf'
                subject = f"{vendor_obj.first_name}{vendor_obj.last_name}  - {vendor_obj.id}-details"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached vendor details - File-{vendor_obj.first_name}{vendor_obj.last_name} .\n--\nRegards,\n", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)
                messages.success(request, 'over view page has been shared via email successfully..!')
                return redirect('view_vendor_details',pk)
    except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('view_vendor_details',pk)
    

def payment_terms_add(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type=='Staff':
            staff_details=StaffDetails.objects.get(login_details=log_details)
            dash_details = CompanyDetails.objects.get(id=staff_details.company.id)

        else:    
            dash_details = CompanyDetails.objects.get(login_details=log_details)        
        if request.method == 'POST':
            terms = request.POST.get('name')
            day = request.POST.get('days')
            ptr = Company_Payment_Term(term_name=terms, days=day, company=dash_details)
            ptr.save()
            payterms_obj = Company_Payment_Term.objects.filter(company=dash_details).values('id', 'term_name')


            payment_list = [{'id': pay_terms['id'], 'name': pay_terms['term_name']} for pay_terms in payterms_obj]
            response_data = {
            "message": "success",
            'payment_list':payment_list,
            }
            return JsonResponse(response_data)

        else:
            return JsonResponse({'error': 'Invalid request'}, status=400)   
            

#---------------------------------------End----------------------------------------------------------------            